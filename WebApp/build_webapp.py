"""
build_webapp.py
================
Generator für den Heidelberg-Haushaltsdaten-Browser (Single-HTML-App).

Liest:  C:\\Users\\D061012\\Desktop\\HD Haushalt\\Daten\\01_HD_Haushaltsdaten.xlsx
        + WebApp\\template.html (Layout / UI)
Schreibt: C:\\Users\\D061012\\Desktop\\HD Haushalt\\WebApp\\index.html

Aufruf: python WebApp/build_webapp.py
"""

from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime
import json
import sys

# ---------------------------------------------------------------------------
# Konfiguration
# ---------------------------------------------------------------------------

BASE      = Path(r"C:\Users\D061012\Desktop\HD Haushalt")
XLSX      = BASE / "Daten" / "01_HD_Haushaltsdaten.xlsx"
OUT_HTML  = BASE / "WebApp" / "index.html"
TEMPLATE_FILE = BASE / "WebApp" / "template.html"

POPULATED_SHEETS = [
    "00_README",
    "01_Eckdaten",
    "02_Teilhaushalte",
    "08_Schulden_Liquiditaet",
    "09_Kennzahlen",
    "13_Mapping_TH_PB",
    "14_Stammdaten_TH",
]

SHEET_META = {
    "00_README":              {"label": "00 · README",            "icon": "📖"},
    "01_Eckdaten":            {"label": "01 · Eckdaten",          "icon": "📊"},
    "02_Teilhaushalte":       {"label": "02 · Teilhaushalte",     "icon": "🏛️"},
    "08_Schulden_Liquiditaet":{"label": "08 · Schulden",          "icon": "💰"},
    "09_Kennzahlen":          {"label": "09 · Kennzahlen",        "icon": "📐"},
    "13_Mapping_TH_PB":       {"label": "13 · Mapping TH⇄PB",     "icon": "🔗"},
    "14_Stammdaten_TH":       {"label": "14 · Stammdaten TH",     "icon": "🗂️"},
}

# ---------------------------------------------------------------------------
# Excel-Reader
# ---------------------------------------------------------------------------

def read_sheet(ws):
    """Liest ein Worksheet zu (headers, rows) als list/list-of-list."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    data_rows = []
    for r in rows[1:]:
        # Komplett leere Zeilen überspringen
        if all(cell is None or (isinstance(cell, str) and cell.strip() == "") for cell in r):
            continue
        data_rows.append(list(r))
    return headers, data_rows


def read_readme(ws):
    """00_README ist Key-Value, kein Tabellenformat."""
    pairs = []
    for r in ws.iter_rows(values_only=True):
        if r and r[0]:
            label = str(r[0])
            value = "" if (len(r) < 2 or r[1] is None) else str(r[1])
            pairs.append({"label": label, "value": value})
    return pairs


def read_kennzahlen(ws):
    """09_Kennzahlen: nur Zeilen mit 4-stelliger Jahreszahl als Daten,
    Rest als Hinweise."""
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) if h is not None else "" for h in rows[0]]
    data_rows = []
    hinweise = []
    for r in rows[1:]:
        if r and r[0] is not None:
            # Versuche, Jahr zu erkennen
            try:
                yr = int(r[0])
                if 2000 <= yr <= 2050:
                    data_rows.append(list(r))
                    continue
            except (ValueError, TypeError):
                pass
            # Sonst: Text-Zeile als Hinweis sammeln
            text = " ".join(str(c) for c in r if c is not None).strip()
            if text:
                hinweise.append(text)
    return headers, data_rows, hinweise


# ---------------------------------------------------------------------------
# Hauptlogik: Daten aufbereiten
# ---------------------------------------------------------------------------

def build_data():
    if not XLSX.exists():
        sys.exit(f"[FEHLER] Excel-Datei nicht gefunden: {XLSX}")
    wb = load_workbook(XLSX, data_only=True)

    sheets = {}
    for name in POPULATED_SHEETS:
        if name not in wb.sheetnames:
            print(f"[WARNUNG] Sheet '{name}' nicht im Workbook — übersprungen.")
            continue
        ws = wb[name]
        if name == "00_README":
            sheets[name] = {
                "type": "kv",
                "pairs": read_readme(ws),
                "label": SHEET_META[name]["label"],
                "icon": SHEET_META[name]["icon"],
            }
        elif name == "09_Kennzahlen":
            headers, data_rows, hinweise = read_kennzahlen(ws)
            sheets[name] = {
                "type": "table",
                "headers": headers,
                "rows": data_rows,
                "hinweise": hinweise,
                "label": SHEET_META[name]["label"],
                "icon": SHEET_META[name]["icon"],
            }
        else:
            headers, data_rows = read_sheet(ws)
            sheets[name] = {
                "type": "table",
                "headers": headers,
                "rows": data_rows,
                "label": SHEET_META[name]["label"],
                "icon": SHEET_META[name]["icon"],
            }

    return {
        "generated": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "sheets": sheets,
    }


# ---------------------------------------------------------------------------
# HTML-Template — wird aus externer Datei geladen
# ---------------------------------------------------------------------------

def load_template():
    if not TEMPLATE_FILE.exists():
        sys.exit(f"[FEHLER] template.html nicht gefunden: {TEMPLATE_FILE}")
    return TEMPLATE_FILE.read_text(encoding="utf-8")


# ---------------------------------------------------------------------------
# Render & write
# ---------------------------------------------------------------------------

def main():
    OUT_HTML.parent.mkdir(parents=True, exist_ok=True)
    print(f"Lese: {XLSX}")
    data = build_data()
    print(f"Sheets geladen: {list(data['sheets'].keys())}")
    for name, sh in data["sheets"].items():
        if sh["type"] == "kv":
            print(f"  {name}: {len(sh['pairs'])} Key-Value-Paare")
        else:
            print(f"  {name}: {len(sh['rows'])} Datenzeilen, {len(sh['headers'])} Spalten")

    template = load_template()
    html = template.replace(
        "__GEN__", data["generated"]
    ).replace(
        "__DATA__", json.dumps(data, ensure_ascii=False, default=str)
    ).replace(
        "__SHEET_ORDER__", json.dumps(POPULATED_SHEETS)
    )

    OUT_HTML.write_text(html, encoding="utf-8")
    size_kb = OUT_HTML.stat().st_size / 1024
    print(f"\n[OK] index.html geschrieben: {OUT_HTML}")
    print(f"     Größe: {size_kb:.1f} KB")


if __name__ == "__main__":
    main()
