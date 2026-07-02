"""
build_vergleichsstaedte_data.py
================================
Baut vergleich-data.js — ein JSON-artiges JS-Modul mit der konsolidierten
Vergleichsstädte-Kennzahlentabelle für den WebApp-Tab "Vergleichsstädte".

Datenquellen:
1. Daten/HD_Kennzahlen_Staedtevergleich.xlsx (Sheets A_Pro-Kopf-Kennzahlen + B_Stueckkosten_Mengen)
2. Daten/01_HD_Haushaltsdaten.xlsx Sheet 11_Vergleichsstaedte (Wegweiser 2023, 6 Städte × 10 KPIs)
3. Daten/01_HD_Haushaltsdaten.xlsx Sheet 09_Kennzahlen (HD-Werte 2023–2026)

Regel:
- HD-Wert wird nur gesetzt, wenn er direkt in einer der drei Quellen steht.
- Städte-Spalte listet konkrete Stadtwerte (falls vorhanden im Kennzahlen-Text/Sheet 11).
- Median/Min/Max/Status kommen aus der Kennzahlen-Datei.

Output: WebApp/vergleich-data.js
"""

from pathlib import Path
from openpyxl import load_workbook
import json, re, sys

BASE = Path(r"C:\Users\D061012\Desktop\HD Haushalt")
MASTER = BASE / "Daten" / "01_HD_Haushaltsdaten.xlsx"
STAEDTE = BASE / "Daten" / "HD_Kennzahlen_Staedtevergleich.xlsx"
OUT = BASE / "WebApp" / "vergleich-data.js"

# ---------------------------------------------------------------------------
# Kategorie-Mapping — gruppiert die 75 Kennzahlen semantisch für den GR-Termin
# ---------------------------------------------------------------------------
CATEGORY_MAP = {
    # A_Pro-Kopf-Kennzahlen — meist mit PB-Bezug
    "PB 11": "Verwaltung & Personal",
    "PB 12": "Sicherheit & Ordnung",
    "PB 126": "Sicherheit & Ordnung",
    "PB 21-24": "Bildung & Schulen",
    "PB 25/27": "Kultur",
    "PB 27": "Kultur",
    "PB 31": "Soziales",
    "PB 36": "Kinder-, Jugend- u. Familienhilfe",
    "PB 42": "Sport",
    "PB 51-52": "Bauen & Wohnen",
    "PB 54": "Verkehr & Straßen",
    "PB 55": "Grün & Umwelt",
    "PB 56": "Grün & Umwelt",
    "PB 57": "Wirtschaft",
    "—": "Rahmenwerte (Land/Bund)",
}

# B_Stueckkosten_Mengen — Bereichs-Feld
BEREICH_TO_CATEGORY = {
    "Schule": "Bildung & Schulen",
    "Jugendhilfe": "Kinder-, Jugend- u. Familienhilfe",
    "Theater": "Kultur",
    "Orchester": "Kultur",
    "Bibliothek": "Kultur",
    "Musikschule": "Kultur",
    "VHS": "Kultur",
    "Bäder": "Sport",
    "ÖPNV": "Verkehr & Straßen",
    "Straßen": "Verkehr & Straßen",
    "Abfall": "Grün & Umwelt",
    "Grünfläche": "Grün & Umwelt",
    "Verwaltung": "Verwaltung & Personal",
    "Bauamt": "Bauen & Wohnen",
    "Friedhof": "Grün & Umwelt",
    "Feuerwehr": "Sicherheit & Ordnung",
    "Kita": "Kinder-, Jugend- u. Familienhilfe",
}

# HD Einwohner
HD_EW = 162960

# ---------------------------------------------------------------------------
# Load HD-Werte aus Master-Excel
# ---------------------------------------------------------------------------

def load_hd_kpis():
    """Extrahiert HD-Werte aus Sheet 11 + Sheet 09."""
    wb = load_workbook(MASTER, data_only=True)
    hd = {}  # kpi_key -> {"value": ..., "unit": ..., "source": ..., "jahr": ...}

    # Sheet 11 — Wegweiser 2023 HD-Zeilen
    ws11 = wb["11_Vergleichsstaedte"]
    for row in ws11.iter_rows(min_row=2, values_only=True):
        if not row[0] or "Heidelberg" not in str(row[0]):
            continue
        stadt, bl, jahr, kpi, wert, einheit, quelle, url, hinweis = row
        if wert is None:
            continue
        # Normalisierter Key
        key = str(kpi).strip().lower()
        hd[key] = {
            "value": wert,
            "unit": einheit,
            "source": f"{quelle} ({jahr})",
            "jahr": jahr,
        }

    # Sheet 09 — HD-Kennzahlen 2023-2026 (nur Zeilen mit Jahr als Integer)
    ws09 = wb["09_Kennzahlen"]
    for row in ws09.iter_rows(min_row=2, values_only=True):
        try:
            jahr = int(row[0])
            if not (2023 <= jahr <= 2029):
                continue
        except (ValueError, TypeError):
            continue
        _, art, buch, nr, benennung, aufgliederung, wert, einheit, plan, quelle, seite, hinweis = row
        if wert is None:
            continue
        # Wir nehmen den 2025-Plan-Wert wenn vorhanden, sonst 2023-Ist
        # Zwei Keys: Benennung + Aufgliederung
        key = f"{benennung}::{aufgliederung}".strip().lower()
        # Nur eintragen wenn noch nicht (2023 hat Vorrang, dann 2025)
        prio = {2023: 1, 2024: 2, 2025: 3, 2026: 4}.get(jahr, 9)
        prev = hd.get(key)
        if prev is None or prev.get("_prio", 99) > prio:
            hd[key] = {
                "value": wert,
                "unit": einheit,
                "source": f"HD Haushalt {jahr} {art} (Sheet 09)",
                "jahr": jahr,
                "_prio": prio,
            }

    return hd


def load_wegweiser_grid():
    """Sheet 11 Wegweiser-Grid — 6 Städte × 10 KPIs (2023)."""
    wb = load_workbook(MASTER, data_only=True)
    ws = wb["11_Vergleichsstaedte"]
    grid = {}  # kpi -> {stadt: wert}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        stadt, bl, jahr, kpi, wert, einheit, quelle, url, hinweis = row
        # Wegweiser-Zeilen (nicht die zwei berechneten HD-Sonderzeilen)
        if "berechnet" in str(kpi).lower() or "Stadtkonzern" in str(kpi):
            continue
        if kpi not in grid:
            grid[kpi] = {"unit": einheit, "cities": {}, "jahr": jahr}
        grid[kpi]["cities"][stadt] = wert
    return grid


# ---------------------------------------------------------------------------
# Load Kennzahlen-Staedtevergleich (75 KPIs)
# ---------------------------------------------------------------------------

def parse_num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return v
    try:
        return float(str(v).replace(",", "."))
    except (ValueError, TypeError):
        return None


def load_sheet_a():
    """Blatt A: 27 Pro-Kopf-Kennzahlen."""
    wb = load_workbook(STAEDTE, data_only=True)
    ws = wb["A_Pro-Kopf-Kennzahlen"]
    kpis = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        # Header-Zeilen (Kategorie-Trenner) haben Nr = None und PB = None
        nr = row[0]
        if not isinstance(nr, (int, float)):
            continue
        pb = row[1] or "—"
        name = row[2]
        bezug = row[3]
        jahr = row[4]
        brutto = parse_num(row[5])
        netto = parse_num(row[6])
        median = parse_num(row[7])
        vmin = parse_num(row[8])
        vmax = parse_num(row[9])
        n = row[10]
        status = row[11]
        beschr = row[12]
        interp = row[13]
        quelle = row[14]
        # Einheit: hier immer €/EW (Blatt-Konvention), teilweise "je Schüler"
        # Wir nutzen brutto oder netto als "wert-referenz"
        wert_ref = brutto if brutto is not None else netto
        category = CATEGORY_MAP.get(str(pb).strip(), "Sonstige")
        kpis.append({
            "sheet": "A",
            "nr": int(nr),
            "category": category,
            "pb": pb,
            "name": name,
            "unit": "€/EW",
            "bezug": bezug,
            "jahr": str(jahr) if jahr else "",
            "brutto": brutto,
            "netto": netto,
            "median": median,
            "min": vmin,
            "max": vmax,
            "n": n,
            "status": status,
            "beschreibung": beschr,
            "interpretation": interp,
            "quelle": quelle,
        })
    return kpis


def load_sheet_b():
    """Blatt B: 47 Stückkosten/Mengen-Kennzahlen."""
    wb = load_workbook(STAEDTE, data_only=True)
    ws = wb["B_Stueckkosten_Mengen"]
    kpis = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        nr = row[0]
        if not isinstance(nr, (int, float)):
            continue
        bereich = row[1]
        name = row[2]
        einheit = row[3]
        bezug = row[4]
        jahr = row[5]
        wert = parse_num(row[6])
        median = parse_num(row[7])
        vmin = parse_num(row[8])
        vmax = parse_num(row[9])
        n = row[10]
        status = row[11]
        beschr = row[12]
        interp = row[13]
        quelle = row[14]
        category = BEREICH_TO_CATEGORY.get(str(bereich).strip(), "Sonstige")
        kpis.append({
            "sheet": "B",
            "nr": int(nr),
            "category": category,
            "bereich": bereich,
            "name": name,
            "unit": einheit,
            "bezug": bezug,
            "jahr": str(jahr) if jahr else "",
            "wert": wert,
            "median": median,
            "min": vmin,
            "max": vmax,
            "n": n,
            "status": status,
            "beschreibung": beschr,
            "interpretation": interp,
            "quelle": quelle,
        })
    return kpis


# ---------------------------------------------------------------------------
# HD-Werte-Matching — heuristisch anhand des KPI-Namens
# ---------------------------------------------------------------------------

def find_hd_value(kpi_name, hd_kpis, wegweiser_grid):
    """Sucht einen HD-Wert für diese Kennzahl. Rückgabe: (wert, unit, quelle) oder None."""
    name_lower = str(kpi_name).lower()

    # 1. Direkter Match aus Sheet 11 Wegweiser
    for wkpi, wdata in wegweiser_grid.items():
        wkpi_l = wkpi.lower()
        # Personalauszahlungen → "Personalaufwand" trifft
        matches = [
            ("personalaufwand", "personalauszahlungen"),
            ("personalauszahlungen", "personalauszahlungen"),
            ("steuereinnahmen", "steuereinnahmen"),
            ("verschuldung kernhaushalte", "verschuldung kernhaushalt"),
        ]
        for name_pat, w_pat in matches:
            if name_pat in name_lower and w_pat in wkpi_l:
                v = wdata["cities"].get("Heidelberg")
                if v is not None:
                    return v, wdata["unit"], f"Wegweiser Kommune {wdata['jahr']} (Sheet 11)"

    # 2. Direktes hd_kpi-Match
    # z.B. "verschuldung kernhaushalt" oder "einwohner"
    for key, val in hd_kpis.items():
        if "verschuldung kernhaushalt" in name_lower and "verschuldung kernhaushalt" in key and "berechnet" not in key:
            return val["value"], val["unit"], val["source"]

    # 3. Grünanlagen 7.2 m²/EW ist im Sheet B direkt gepflegt (Zeile 45)
    # Wird durch normale wert-Spalte abgedeckt.

    return None, None, None


# ---------------------------------------------------------------------------
# Zusammenführung + Ausgabe
# ---------------------------------------------------------------------------

def merge_all():
    hd = load_hd_kpis()
    wegweiser = load_wegweiser_grid()
    sheet_a = load_sheet_a()
    sheet_b = load_sheet_b()

    rows = []

    # === Wegweiser 6-Städte-Grid als eigene Sektion (10 KPIs) ===
    for wkpi, wdata in wegweiser.items():
        if wkpi == "Einwohner":
            continue  # Meta-Info, nicht KPI
        cities = wdata["cities"]
        hd_val = cities.get("Heidelberg")
        others = {k: v for k, v in cities.items() if k != "Heidelberg"}
        vals = [v for v in others.values() if isinstance(v, (int, float))]
        median = None
        vmin = None
        vmax = None
        if len(vals) >= 3:
            sorted_v = sorted(vals)
            n = len(sorted_v)
            median = sorted_v[n//2] if n % 2 else (sorted_v[n//2-1] + sorted_v[n//2]) / 2
            vmin = min(vals)
            vmax = max(vals)
        rows.append({
            "category": "Kernhaushalt (Wegweiser 2023, BW+HE)",
            "name": wkpi,
            "unit": wdata["unit"],
            "hd_value": hd_val,
            "median": median,
            "min": vmin,
            "max": vmax,
            "n": len(vals),
            "status": "6-Städte-Vergleich (Wegweiser)",
            "cities": {k: v for k, v in cities.items() if k != "Heidelberg"},
            "beschreibung": f"Bertelsmann Wegweiser Kommune {wdata['jahr']} — 5 vergleichbare Stadtkreise (KA, MA, FR, TÜ, DA).",
            "interpretation": f"Median berechnet über die 5 Vergleichsstädte (ohne HD). Direkt vergleichbar (standardisierte Länderstatistik).",
            "quelle": "Bertelsmann Wegweiser Kommune (Finanzbericht 2023)",
            "jahr": str(wdata["jahr"]),
            "sheet": "M",  # Master
        })

    # === Blatt A – 27 Pro-Kopf-Kennzahlen ===
    for k in sheet_a:
        hd_val, hd_unit, hd_src = find_hd_value(k["name"], hd, wegweiser)
        rows.append({
            "category": k["category"],
            "name": k["name"],
            "unit": k["unit"],
            "hd_value": hd_val,
            "hd_source": hd_src,
            "median": k["median"],
            "min": k["min"],
            "max": k["max"],
            "n": k["n"],
            "status": k["status"],
            "cities": {},  # Blatt A hat keine Stadt-Aufspaltung
            "beschreibung": k["beschreibung"],
            "interpretation": k["interpretation"],
            "quelle": k["quelle"],
            "jahr": k["jahr"],
            "bezug": k["bezug"],
            "brutto": k["brutto"],
            "netto": k["netto"],
            "pb": k["pb"],
            "sheet": "A",
        })

    # === Blatt B – 47 Stückkosten/Mengen ===
    # Grünanlagen HD-Wert (7,2 m²/EW) ist als Einzelwert in Zeile 45 im Sheet B
    # → Wir sammeln alle Grünanlagen-Einzelstädte in EINE Zeile (KPI 47)
    gruen_cities = {}
    gruen_hd = None
    for k in sheet_b:
        if "Grünanlagenfläche" in str(k["name"]) and "Richtwert" not in str(k["name"]):
            m = re.search(r"([A-Za-zäöüÄÖÜß ]+)\s*\((BW|HE|NRW)\)", str(k["bezug"]))
            stadt = m.group(1).strip() if m else str(k["bezug"])
            if "Heidelberg" in stadt:
                gruen_hd = k["wert"]
            else:
                gruen_cities[stadt] = k["wert"]

    for k in sheet_b:
        # Grünanlage-Einzelstädte werden übersprungen (wir haben Aggregat-Zeile)
        if "Grünanlagenfläche" in str(k["name"]) and "Richtwert" not in str(k["name"]):
            continue

        # Bei "→ Richtwert Grünanlagenfläche" (KPI 47): mit Städten anreichern
        cities = {}
        hd_val = None
        if "Grünanlagenfläche" in str(k["name"]) and "Richtwert" in str(k["name"]):
            cities = gruen_cities
            hd_val = gruen_hd
        else:
            hd_val, _, _ = find_hd_value(k["name"], hd, wegweiser)

        rows.append({
            "category": k["category"],
            "name": k["name"],
            "unit": k["unit"],
            "hd_value": hd_val,
            "hd_source": None,
            "median": k["median"] if k["median"] is not None else k["wert"],
            "min": k["min"],
            "max": k["max"],
            "n": k["n"],
            "status": k["status"],
            "cities": cities,
            "beschreibung": k["beschreibung"],
            "interpretation": k["interpretation"],
            "quelle": k["quelle"],
            "jahr": k["jahr"],
            "bezug": k["bezug"],
            "bereich": k["bereich"],
            "sheet": "B",
        })

    # Kategorien sortieren
    CAT_ORDER = [
        "Kernhaushalt (Wegweiser 2023, BW+HE)",
        "Rahmenwerte (Land/Bund)",
        "Verwaltung & Personal",
        "Bildung & Schulen",
        "Kinder-, Jugend- u. Familienhilfe",
        "Soziales",
        "Kultur",
        "Sport",
        "Sicherheit & Ordnung",
        "Bauen & Wohnen",
        "Verkehr & Straßen",
        "Grün & Umwelt",
        "Wirtschaft",
        "Sonstige",
    ]
    def cat_sort(r):
        try:
            i = CAT_ORDER.index(r["category"])
        except ValueError:
            i = 99
        return (i, r["name"])
    rows.sort(key=cat_sort)

    return {
        "hd_ew": HD_EW,
        "generated": None,  # wird im build_webapp.py gesetzt
        "categories": CAT_ORDER,
        "rows": rows,
    }


def main():
    data = merge_all()
    js = "// Auto-generated by build_vergleichsstaedte_data.py — nicht editieren\n"
    js += "window.VERGLEICH_DATA = " + json.dumps(data, ensure_ascii=False, default=str, indent=2) + ";\n"
    OUT.write_text(js, encoding="utf-8")
    n_rows = len(data["rows"])
    n_with_hd = sum(1 for r in data["rows"] if r.get("hd_value") is not None)
    n_with_cities = sum(1 for r in data["rows"] if r.get("cities"))
    size_kb = OUT.stat().st_size / 1024
    print(f"[OK] {OUT}")
    print(f"     Rows: {n_rows} (davon {n_with_hd} mit HD-Wert, {n_with_cities} mit Städteliste)")
    print(f"     Kategorien: {len(set(r['category'] for r in data['rows']))}")
    print(f"     Größe: {size_kb:.1f} KB")


if __name__ == "__main__":
    main()
